import streamlit as st
import pandas as pd
from PIL import Image
import openpyxl

st.set_page_config(page_title = "BIO-DATA OF STUDENTS", page_icon=":smile:")
title = st.container()
Create = st.container()
Update = st.container()
Read = st.container()

with title:
    st.header("BIO-DATA OF STUDENTS")
    st.write("HERE YOU CAN INFORMATION ABOUT A STUDENT!")

df = pd.read_excel('C:\sanjai\Chitra English\Students.xlsx')

rows = 0
for row in df.iterrows():
    rows = rows + 1

with Create:
    rows = rows + 2
    st.header("CREATE AN ID!")
    name = st.text_input("NAME: ").strip().upper()
    admission_no = st.text_input("ADMISSION NO: ").strip().upper()
    clas = st.text_input("CLASS: ").strip().upper()
    section = st.text_input("SECTION: ").strip().upper()
    roll_no = st.text_input("ROLL NO: ").strip().upper()
    if name != "":
        if admission_no != "":
            if clas != "":
                if section != "":
                    if roll_no != "":
                        df2 = openpyxl.load_workbook('C:\sanjai\Chitra English\Students.xlsx')
                        sheet = df2["Sheet1"]
                        sheet.cell(row = rows, column = 1, value = str(name))
                        sheet.cell(row=rows, column=2, value = str(admission_no))
                        sheet.cell(row=rows, column=3, value = str(clas))
                        sheet.cell(row=rows, column=4, value = str(section))
                        sheet.cell(row=rows, column=5, value = str(roll_no))
                        df2.save('C:\sanjai\Chitra English\Students.xlsx')
    else:
        pass

with Update:
    st.header("UPDATE!")
    st.write("HERE YOU CAN UPDATE YOUR PERSONAL DETAILS!")
    rows = rows - 2
    NAme = st.text_input("Name: ").strip().upper()
    heading = st.text_input("HEADING: ").strip().upper()
    for i in range(0, rows):
        if NAme == df.iloc[i, 0]:
            if heading == "ADMISSION NO":
                new = st.text_input("NEW INPUT: ")
                df.iloc[i, 1] = new
            elif heading == "CLASS":
                new = st.text_input("NEW INPUT: ")
                df.iloc[i, 2] = new
            elif heading == "SECTION":
                new = st.text_input("NEW INPUT: ")
                df.iloc[i, 3] = new
            elif heading == "ROLL NO":
                new = st.text_input("NEW INPUT: ")
                df.iloc[i, 4] = new
            else:
                pass

with Read:
    st.header("SHOW AN ID!")
    Admission_no = st.text_input("Enter Admission no: ").strip().upper()
    for i in range(0, rows):
        if Admission_no == df.iloc[i, 1]:
            Name = df.iloc[i, 0]
            Admission_no = df.iloc[i, 1]
            Class = df.iloc[i, 2]
            Section = df.iloc[i, 3]
            Roll_no = df.iloc[i, 4]
            Photo_url = df.iloc[i, 5]
        else:
            pass

    try:
        if Name != "":
            if Admission_no != "":
                if Class != "":
                    if Section != "":
                        if Roll_no != "":
                            st.write("NAME: " + str(Name))
                            st.write("ADMISSION NO: " + str(Admission_no))
                            st.write("CLASS: " + str(Class))
                            st.write("SECTION: " + str(Section))
                            st.write("ROLL NO: " + str(Roll_no))
                            photo = Image.open(Photo_url)
                            st.image(photo, caption = Name, width=500, use_column_width=200)
    except:
        pass

st.dataframe(df[["NAME", "ADMISSION NO", "CLASS", "SECTION", "ROLL NO"]])




